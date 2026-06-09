from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──────────────────────────────────────────────────────────
blue_fill = PatternFill('solid', fgColor='0052D9')
light_blue_fill = PatternFill('solid', fgColor='D6E4FF')
light_gray_fill = PatternFill('solid', fgColor='F5F6F7')
title_font = Font(name='Arial', bold=True, color='0052D9', size=16)
section_font = Font(name='Arial', bold=True, color='0052D9', size=13)
subsection_font = Font(name='Arial', bold=True, color='1D1D1F', size=11)
normal_font = Font(name='Arial', color='1D1D1F', size=10)
bold_font = Font(name='Arial', bold=True, color='1D1D1F', size=10)
hint_font = Font(name='Arial', color='B0B0B2', size=10, italic=True)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
thin_border = Border(
    left=Side(style='thin', color='E5E6EB'),
    right=Side(style='thin', color='E5E6EB'),
    top=Side(style='thin', color='E5E6EB'),
    bottom=Side(style='thin', color='E5E6EB'),
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)


def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = blue_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border


def write_row(ws, row, data, font_style=None, alignment=None, fill=None, col_start=1):
    if font_style is None: font_style = normal_font
    if alignment is None: alignment = left_align
    for i, val in enumerate(data):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font_style
        cell.alignment = alignment
        cell.border = thin_border
        if fill: cell.fill = fill


def merge_and_fill(ws, row, col_start, col_end, value, font_style=None, alignment=None):
    if font_style is None: font_style = normal_font
    if alignment is None: alignment = left_align
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = font_style
    cell.alignment = alignment
    cell.border = thin_border
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border


def add_title(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = light_gray_fill
        ws.cell(row=row, column=c).border = thin_border
    return row + 2


def add_section_title(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.alignment = left_align
    return row + 1


def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =====================================================================
# Sheet 1: 报告基本信息
# =====================================================================
ws = wb.active
ws.title = '1.报告基本信息'
ncols = 6
set_cols(ws, [18, 18, 18, 18, 18, 18])
row = add_title(ws, 1, '一、报告基本信息', ncols)

fields = [
    ('评估部门', ''), ('部门负责人', ''), ('主要评估人', ''),
    ('参与试用人员', ''), ('试用周期', '    年   月   日 至    年   月   日'),
    ('报告提交日期', '    年   月   日'),
]
for label, value in fields:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=1).fill = light_blue_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    ws.cell(row=row, column=2, value=value).font = normal_font
    ws.cell(row=row, column=2).alignment = left_align
    ws.cell(row=row, column=2).border = thin_border
    for c in range(3, ncols + 1):
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1

# =====================================================================
# Sheet 2: 总体评估概述
# =====================================================================
ws = wb.create_sheet('2.总体评估概述')
set_cols(ws, [18] + [18]*5)
row = add_title(ws, 1, '二、总体评估概述', ncols)

merge_and_fill(ws, row, 1, ncols, '2.1 平台整体评价', subsection_font)
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=ncols)
cell = ws.cell(row=row, column=1, value='（请用100-200字概括本部门对智控平台的整体印象）')
cell.font = hint_font
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
cell.border = thin_border
for c in range(2, ncols + 1):
    for rr in range(row, row + 4):
        ws.cell(row=rr, column=c).border = thin_border
for rr in range(row, row + 4):
    ws.row_dimensions[rr].height = 24
row += 5

merge_and_fill(ws, row, 1, ncols, '2.2 核心结论', subsection_font)
row += 1
for item in [
    '已满足的核心业务需求：',
    '最迫切需要解决的问题：',
    '平台整体评分：□ 优秀（90-100分）  □ 良好（80-89分）  □ 一般（70-79分）  □ 较差（60-69分）  □ 很差（60分以下）',
]:
    merge_and_fill(ws, row, 1, ncols, item, normal_font)
    row += 1

# =====================================================================
# Sheet 3: 业务功能匹配度详细评估
# =====================================================================
ws = wb.create_sheet('3.业务功能匹配度评估')
set_cols(ws, [22, 26, 20, 24, 22, 24, 16])
ncols3 = 7
row = add_title(ws, 1, '三、业务功能匹配度详细评估', ncols3)

headers = ['业务管理线条', '具体业务需求描述', '平台实现情况', '已实现功能说明及优势',
           '存在的问题/不足', '待开发功能详细描述', '业务影响程度']
for i, h in enumerate(headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, ncols3)
row += 1

modules = [
    ('模块一：__________管理', 3),
    ('模块二：__________管理', 3),
    ('模块三：__________管理', 2),
]
item_idx = 1
for mod_name, count in modules:
    start_r = row
    for j in range(count):
        label = f'{item_idx}.{j + 1} __________业务流程'
        item_idx += 1
        write_row(ws, row, [label, '', '□ 完全实现  □ 部分实现  □ 未实现', '', '', '', '□ 高  □ 中  □ 低'])
        ws.row_dimensions[row].height = 36
        row += 1
    end_r = row - 1
    ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
    cell = ws.cell(row=start_r, column=1, value=mod_name)
    cell.font = bold_font
    cell.fill = light_blue_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    for r in range(start_r, end_r + 1):
        for c in range(1, ncols3 + 1):
            ws.cell(row=r, column=c).border = thin_border

# =====================================================================
# Sheet 4: 系统性能评估
# =====================================================================
ws = wb.create_sheet('4.系统性能评估')
set_cols(ws, [28, 30, 16, 30])
ncols4 = 4
row = add_title(ws, 1, '四、系统性能与易用性评估 — 4.1 系统性能评估', ncols4)

for i, h in enumerate(['评估指标', '实际情况', '评分(1-10分)', '问题描述'], 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, ncols4)
row += 1

for item in ['页面加载速度', '数据查询响应速度', '系统稳定性（崩溃/卡顿频率）', '并发使用支持能力', '大数据量处理能力']:
    write_row(ws, row, [item, '', '', ''])
    ws.row_dimensions[row].height = 28
    row += 1

# =====================================================================
# Sheet 5: 系统易用性评估
# =====================================================================
ws = wb.create_sheet('5.系统易用性评估')
set_cols(ws, [28, 30, 16, 30])
row = add_title(ws, 1, '四、系统性能与易用性评估 — 4.2 系统易用性评估', ncols4)

for i, h in enumerate(['评估指标', '实际情况', '评分(1-10分)', '问题描述'], 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, ncols4)
row += 1

for item in ['界面设计合理性', '操作流程便捷性', '功能导航清晰度', '错误提示明确性', '帮助文档完整性']:
    write_row(ws, row, [item, '', '', ''])
    ws.row_dimensions[row].height = 28
    row += 1

# =====================================================================
# Sheet 6: 数据管理评估
# =====================================================================
ws = wb.create_sheet('6.数据管理评估')
set_cols(ws, [26] + [20]*5)
ncols6 = 6
row = add_title(ws, 1, '五、数据管理与安全评估 — 5.1 数据管理评估', ncols6)

items = [
    '数据准确性：□ 很高  □ 较高  □ 一般  □ 较低  □ 很低',
    '数据完整性：□ 很高  □ 较高  □ 一般  □ 较低  □ 很低',
    '数据查询统计功能：□ 很强  □ 较强  □ 一般  □ 较弱  □ 很弱',
    '数据导出功能：□ 很强  □ 较强  □ 一般  □ 较弱  □ 很弱',
]
for item in items:
    merge_and_fill(ws, row, 1, ncols6, item, normal_font)
    row += 1
    merge_and_fill(ws, row, 1, ncols6, '    问题描述：', bold_font)
    row += 1
    write_row(ws, row, [''] * ncols6)
    ws.row_dimensions[row].height = 32
    row += 1

# =====================================================================
# Sheet 7: 权限与安全评估
# =====================================================================
ws = wb.create_sheet('7.权限与安全评估')
set_cols(ws, [26] + [20]*5)
row = add_title(ws, 1, '五、数据管理与安全评估 — 5.2 权限与安全评估', ncols6)

items = [
    '权限管理精细度：□ 很高  □ 较高  □ 一般  □ 较低  □ 很低',
    '数据访问安全性：□ 很高  □ 较高  □ 一般  □ 较低  □ 很低',
    '操作日志完整性：□ 很高  □ 较高  □ 一般  □ 较低  □ 很低',
]
for item in items:
    merge_and_fill(ws, row, 1, ncols6, item, normal_font)
    row += 1
    merge_and_fill(ws, row, 1, ncols6, '    问题描述：', bold_font)
    row += 1
    write_row(ws, row, [''] * ncols6)
    ws.row_dimensions[row].height = 32
    row += 1

# =====================================================================
# Sheet 8: 集成与扩展性评估
# =====================================================================
ws = wb.create_sheet('8.集成与扩展性评估')
set_cols(ws, [26] + [20]*5)
row = add_title(ws, 1, '六、集成与扩展性评估', ncols6)

entries = [
    ('与本部门现有其他系统的集成情况：', '□ 已完全集成  □ 部分集成  □ 未集成'),
    ('自定义功能支持能力：', '□ 很强  □ 较强  □ 一般  □ 较弱  □ 很弱'),
    ('业务流程调整灵活性：', '□ 很强  □ 较强  □ 一般  □ 较弱  □ 很弱'),
]
for label, options in entries:
    merge_and_fill(ws, row, 1, ncols6, label + options, normal_font)
    row += 1
    merge_and_fill(ws, row, 1, ncols6, '    具体情况：', bold_font)
    row += 1
    write_row(ws, row, [''] * ncols6)
    ws.row_dimensions[row].height = 32
    row += 1

# =====================================================================
# Sheet 9: 培训与技术支持评估
# =====================================================================
ws = wb.create_sheet('9.培训与技术支持评估')
set_cols(ws, [26] + [20]*5)
row = add_title(ws, 1, '七、培训与技术支持评估', ncols6)

items = [
    '平台培训效果：□ 很好  □ 较好  □ 一般  □ 较差  □ 很差',
    '技术问题响应速度：□ 很快  □ 较快  □ 一般  □ 较慢  □ 很慢',
    '问题解决能力：□ 很强  □ 较强  □ 一般  □ 较弱  □ 很弱',
]
for item in items:
    merge_and_fill(ws, row, 1, ncols6, item, normal_font)
    row += 1
    merge_and_fill(ws, row, 1, ncols6, '    问题描述：', bold_font)
    row += 1
    write_row(ws, row, [''] * ncols6)
    ws.row_dimensions[row].height = 32
    row += 1

# =====================================================================
# Sheet 10: 待开发功能优先级建议
# =====================================================================
ws = wb.create_sheet('10.待开发功能优先级建议')
set_cols(ws, [16, 24, 22, 36, 18])
ncols10 = 5
row = add_title(ws, 1, '八、待开发功能优先级建议', ncols10)

for i, h in enumerate(['优先级', '待开发功能名称', '所属业务模块', '理由说明', '期望完成时间'], 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, ncols10)
row += 1

for pd in [
    ('P0（最高）', '', '', '', ''),
    ('P0（最高）', '', '', '', ''),
    ('P1（高）', '', '', '', ''),
    ('P1（高）', '', '', '', ''),
    ('P2（中）', '', '', '', ''),
    ('P3（低）', '', '', '', ''),
]:
    write_row(ws, row, pd)
    ws.row_dimensions[row].height = 28
    row += 1

# =====================================================================
# Sheet 11: 其他意见与建议
# =====================================================================
ws = wb.create_sheet('11.其他意见与建议')
set_cols(ws, [18] + [18]*5)
row = add_title(ws, 1, '九、其他意见与建议', ncols6)

ws.merge_cells(start_row=row, start_column=1, end_row=row + 5, end_column=ncols6)
cell = ws.cell(row=row, column=1, value='（请填写上述未覆盖的其他意见和建议）')
cell.font = hint_font
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
cell.border = thin_border
for c in range(2, ncols6 + 1):
    for rr in range(row, row + 6):
        ws.cell(row=rr, column=c).border = thin_border
for rr in range(row, row + 6):
    ws.row_dimensions[rr].height = 24

# =====================================================================
# Sheet 12: 部门确认
# =====================================================================
ws = wb.create_sheet('12.部门确认')
set_cols(ws, [14, 16, 14, 16, 14, 16, 14, 16])
ncols12 = 8
row = add_title(ws, 1, '十、部门确认', ncols12)

conf_headers = ['评估人签字', '', '日期', '', '部门负责人签字', '', '日期', '']
for i, h in enumerate(conf_headers, 1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = thin_border
    if i % 2 == 1:
        cell.fill = light_blue_fill
row += 1
write_row(ws, row, [''] * ncols12, alignment=center_align)
ws.row_dimensions[row].height = 32

# =====================================================================
# Sheet 13: 填写指导
# =====================================================================
ws = wb.create_sheet('13.填写指导')
set_cols(ws, [20] + [24]*5)
ncols13 = 6
row = add_title(ws, 1, '填写指导说明', ncols13)

row = add_section_title(ws, row, '一、评估要求', ncols13)
for req in [
    '1. 请各部门组织本部门所有使用智控平台的人员参与评估',
    '2. 确保评估结果全面、客观、真实反映平台使用情况',
    '3. 重点关注平台与本部门实际业务管理线条的匹配度',
    '4. 对于待开发功能，请尽可能详细描述具体需求和业务场景',
]:
    merge_and_fill(ws, row, 1, ncols13, req, normal_font)
    row += 1
row += 1

row = add_section_title(ws, row, '二、核心部分填写说明', ncols13)
merge_and_fill(ws, row, 1, ncols13, '第三部分"业务功能匹配度详细评估"', subsection_font)
row += 1

for i, h in enumerate(['填写项目', '填写说明'], 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, 2)
row += 1

guidance = [
    ('业务管理线条', '请严格按照本部门实际的业务管理架构和流程填写'),
    ('具体业务需求描述', '详细描述该业务环节需要系统完成的具体功能'),
    ('平台实现情况', '完全实现：平台功能完全满足业务需求\n部分实现：平台功能部分满足，需要人工辅助\n未实现：平台没有该功能'),
    ('业务影响程度', '高：功能缺失导致核心业务无法开展\n中：功能缺失导致业务流程不顺畅\n低：功能缺失对业务影响较小'),
]
for item, desc in guidance:
    write_row(ws, row, [item, desc], alignment=top_left_align)
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=1).fill = light_blue_fill
    ws.row_dimensions[row].height = 50 if '\n' in desc else 28
    row += 1
row += 1

merge_and_fill(ws, row, 1, ncols13, '第八部分"待开发功能优先级建议"', subsection_font)
row += 1

for i, h in enumerate(['优先级级别', '含义说明'], 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, 2)
row += 1

for level, desc in [
    ('P0（最高）', '阻塞核心业务开展，必须立即开发'),
    ('P1（高）', '严重影响业务效率，需要尽快开发'),
    ('P2（中）', '影响部分业务流程，可在后续版本中开发'),
    ('P3（低）', '优化提升类需求，可根据资源情况安排'),
]:
    write_row(ws, row, [level, desc])
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=1).fill = light_blue_fill
    row += 1
row += 1

row = add_section_title(ws, row, '三、提交要求', ncols13)
for item in [
    '请于 ____年__月__日 前将本报告提交至平台项目组',
    '提交方式：________________________',
    '如有疑问，请联系：____________（电话：______________，邮箱：______________）',
]:
    merge_and_fill(ws, row, 1, ncols13, item, normal_font)
    row += 1

# =====================================================================
# Sheet 14: 示例填写
# =====================================================================
ws = wb.create_sheet('14.示例填写')
set_cols(ws, [22, 26, 20, 24, 22, 24, 16])
ncols14 = 7
row = add_title(ws, 1, '填写示例（仅供参考）', ncols14)

row = add_section_title(ws, row, '三、业务功能匹配度详细评估（示例）', ncols14)

for i, h in enumerate(headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, ncols14)
row += 1

example_data = [
    ('模块一：项目进度管理', [
        ('1.1 项目计划编制', '能够创建多级项目计划，设置任务依赖关系和里程碑',
         '□ 部分实现', '支持创建任务和设置开始结束时间',
         '不支持任务依赖关系和里程碑设置', '增加任务依赖关系管理功能，支持里程碑标记', '□ 高'),
        ('1.2 进度填报与更新', '支持按周填报任务完成情况，自动计算项目整体进度',
         '□ 完全实现', '填报界面友好，自动计算进度百分比',
         '暂无明显问题', '', '□ 低'),
        ('1.3 进度报表生成', '能够按周/月生成项目进度报表并导出Excel',
         '□ 部分实现', '支持导出Excel，但报表格式固定',
         '无法自定义报表格式和字段', '增加报表自定义功能，支持用户选择导出字段', '□ 中'),
    ]),
    ('模块二：合同管理', [
        ('2.1 合同信息录入', '支持录入合同基本信息、附件上传和版本管理',
         '□ 完全实现', '附件上传方便，支持多版本管理',
         '暂无明显问题', '', '□ 低'),
        ('2.2 合同审批流程', '支持合同审批流程的自定义配置和流转',
         '□ 未实现', '',
         '平台没有合同审批功能', '开发合同审批模块，支持流程自定义和电子签章', '□ 高'),
    ]),
]

for mod_name, subs in example_data:
    start_r = row
    for j, sub in enumerate(subs):
        write_row(ws, row, list(sub), alignment=top_left_align)
        ws.row_dimensions[row].height = 55
        row += 1
    end_r = row - 1
    ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
    cell = ws.cell(row=start_r, column=1, value=mod_name)
    cell.font = bold_font
    cell.fill = light_blue_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    for r in range(start_r, end_r + 1):
        for c in range(1, ncols14 + 1):
            ws.cell(row=r, column=c).border = thin_border

row += 1
row = add_section_title(ws, row, '八、待开发功能优先级建议（示例）', ncols14)

for i, h in enumerate(['优先级', '待开发功能名称', '所属业务模块', '理由说明', '期望完成时间'], 1):
    ws.cell(row=row, column=i, value=h)
style_header(ws, row, 1, 5)
row += 1

for pd in [
    ('P0（最高）', '合同审批流程', '合同管理', '目前合同审批仍需线下进行，严重影响工作效率', '2026年6月30日'),
    ('P0（最高）', '任务依赖关系管理', '项目进度管理', '无法准确反映项目任务之间的逻辑关系', '2026年6月30日'),
    ('P1（高）', '报表自定义功能', '项目进度管理', '固定格式报表无法满足不同项目的管理需求', '2026年7月31日'),
    ('P2（中）', '移动端进度填报', '项目进度管理', '现场人员无法及时填报进度信息', '2026年8月31日'),
    ('P3（低）', '进度预警功能', '项目进度管理', '自动提醒即将延期的任务', '2026年9月30日'),
]:
    write_row(ws, row, pd, alignment=left_align)
    ws.row_dimensions[row].height = 28
    row += 1

# ── Print settings for all sheets ──
for ws in wb.worksheets:
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9

# ── Save ──
output = r'D:\code\file_sys\frontend\工程智控平台试用评估报告.xlsx'
wb.save(output)
print(f'Done: {output}')
